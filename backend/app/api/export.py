# ファイル: backend/app/api/export.py

from fastapi import APIRouter, Depends, HTTPException, status, Response
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from uuid import UUID
import json
import csv
import io
from datetime import datetime
from typing import Dict, Any

from app.database import get_db
from app.models.test_result import TestResult
from app.models.question import Question
from app.utils.athlete_type_algorithm import analyze_athlete_type

# Import libraries for advanced exports
try:
    import pandas as pd
    import matplotlib.pyplot as plt
    import seaborn as sns
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, Image
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.units import inch
    from reportlab.lib import colors
    from reportlab.graphics.shapes import Drawing
    from reportlab.graphics.charts.piecharts import Pie
    from reportlab.graphics.charts.barcharts import VerticalBarChart
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.chart import BarChart, PieChart, Reference
    import xlsxwriter
except ImportError as e:
    print(f"Export dependencies not installed: {e}")
    print("Run: pip install pandas matplotlib seaborn reportlab openpyxl xlsxwriter")

router = APIRouter()


@router.get("/pdf/{result_id}")
async def export_pdf(
    result_id: UUID,
    db: Session = Depends(get_db)
):
    """PDF形式でテスト結果をエクスポート"""
    try:
        # テスト結果を取得
        test_result = db.query(TestResult).filter(
            TestResult.result_id == result_id
        ).first()
        
        if not test_result:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Test result not found"
            )
        
        # PDF生成
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=A4)
        styles = getSampleStyleSheet()
        story = []
        
        # タイトル
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=24,
            spaceAfter=30,
            alignment=1,  # CENTER
            textColor=colors.HexColor('#667eea')
        )
        
        story.append(Paragraph("スポーツマンシップメンタルテスト結果", title_style))
        story.append(Spacer(1, 20))
        
        # 基本情報
        info_data = [
            ['テスト実施日', test_result.created_date.strftime('%Y年%m月%d日')],
            ['対象', get_target_name(test_result.target_selection)],
            ['結果ID', str(test_result.result_id)[:8] + '...']
        ]
        
        info_table = Table(info_data, colWidths=[2*inch, 3*inch])
        info_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, -1), colors.HexColor('#f8f9fa')),
            ('TEXTCOLOR', (0, 0), (-1, -1), colors.black),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 12),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(info_table)
        story.append(Spacer(1, 30))
        
        # スコアサマリー
        story.append(Paragraph("スコアサマリー", styles['Heading2']))
        story.append(Spacer(1, 12))
        
        # スポーツマンシップスコア
        sportsmanship_scores = [
            test_result.courage or 0,
            test_result.resilience or 0,
            test_result.cooperation or 0,
            test_result.natural_acceptance or 0,
            test_result.non_rationality or 0
        ]
        sportsmanship_total = sum(sportsmanship_scores)
        
        # アスリートマインドスコア
        athlete_mind_scores = [
            test_result.introspection or 0,
            test_result.self_control or 0,
            test_result.devotion or 0,
            test_result.intuition or 0,
            test_result.sensitivity or 0,
            test_result.steadiness or 0,
            test_result.comparison or 0,
            test_result.result or 0,
            test_result.assertion or 0,
            test_result.commitment or 0
        ]
        athlete_mind_total = sum(athlete_mind_scores)
        
        # 自己肯定感スコア
        self_affirmation_scores = [
            test_result.self_determination or 0,
            test_result.self_acceptance or 0,
            test_result.self_worth or 0,
            test_result.self_efficacy or 0
        ]
        self_affirmation_total = sum(self_affirmation_scores)
        
        score_data = [
            ['カテゴリ', '合計スコア', '平均スコア', '評価'],
            ['スポーツマンシップ', str(sportsmanship_total), f"{sportsmanship_total/5:.1f}", get_grade(sportsmanship_total/5, 10)],
            ['アスリートマインド', str(athlete_mind_total), f"{athlete_mind_total/10:.1f}", get_grade(athlete_mind_total/10, 50)],
            ['自己肯定感', str(self_affirmation_total), f"{self_affirmation_total/4:.1f}", get_grade(self_affirmation_total/4, 50)]
        ]
        
        score_table = Table(score_data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1*inch])
        score_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 12),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(score_table)
        story.append(Spacer(1, 30))
        
        # アスリートタイプ分析
        try:
            athlete_mind_data = {
                'commitment': test_result.commitment or 0,
                'results_oriented': test_result.result or 0,
                'steady': test_result.steadiness or 0,
                'devoted': test_result.devotion or 0,
                'self_control': test_result.self_control or 0,
                'assertive': test_result.assertion or 0,
                'sensitive': test_result.sensitivity or 0,
                'intuitive': test_result.intuition or 0,
                'introspective': test_result.introspection or 0,
                'comparative': test_result.comparison or 0
            }
            
            athlete_type = analyze_athlete_type(athlete_mind_data, test_result.target_selection)
            
            story.append(Paragraph("アスリートタイプ判定", styles['Heading2']))
            story.append(Spacer(1, 12))
            story.append(Paragraph(f"<b>{athlete_type['type_name']}</b>", styles['Normal']))
            story.append(Spacer(1, 6))
            story.append(Paragraph(athlete_type['description'], styles['Normal']))
            
        except Exception as e:
            print(f"Athlete type analysis failed: {e}")
        
        # PDFビルド
        doc.build(story)
        buffer.seek(0)
        
        return StreamingResponse(
            io.BytesIO(buffer.read()),
            media_type="application/pdf",
            headers={"Content-Disposition": f"attachment; filename=test_result_{result_id}.pdf"}
        )
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"PDF export failed: {str(e)}"
        )


@router.get("/excel/{result_id}")
async def export_excel(
    result_id: UUID,
    db: Session = Depends(get_db)
):
    """Excel形式でテスト結果をエクスポート"""
    try:
        # テスト結果を取得
        test_result = db.query(TestResult).filter(
            TestResult.result_id == result_id
        ).first()
        
        if not test_result:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Test result not found"
            )
        
        # Excelファイル作成
        buffer = io.BytesIO()
        workbook = openpyxl.Workbook()
        
        # サマリーシート
        ws_summary = workbook.active
        ws_summary.title = "サマリー"
        
        # ヘッダースタイル
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center")
        
        # 基本情報
        ws_summary['A1'] = "スポーツマンシップメンタルテスト結果"
        ws_summary['A1'].font = Font(size=16, bold=True)
        ws_summary.merge_cells('A1:D1')
        
        ws_summary['A3'] = "テスト実施日"
        ws_summary['B3'] = test_result.created_date.strftime('%Y年%m月%d日')
        ws_summary['A4'] = "対象"
        ws_summary['B4'] = get_target_name(test_result.target_selection)
        ws_summary['A5'] = "結果ID"
        ws_summary['B5'] = str(test_result.result_id)
        
        # スコア詳細シート
        ws_scores = workbook.create_sheet("詳細スコア")
        
        # スポーツマンシップ
        ws_scores['A1'] = "スポーツマンシップ"
        ws_scores['A1'].font = header_font
        ws_scores['A1'].fill = header_fill
        
        sportsmanship_data = [
            ['項目', 'スコア'],
            ['勇気', test_result.courage or 0],
            ['精神的強さ', test_result.resilience or 0],
            ['協調性', test_result.cooperation or 0],
            ['自然受容', test_result.natural_acceptance or 0],
            ['非合理性', test_result.non_rationality or 0]
        ]
        
        for i, row in enumerate(sportsmanship_data):
            for j, value in enumerate(row):
                cell = ws_scores.cell(row=i+2, column=j+1, value=value)
                if i == 0:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
        
        # アスリートマインド
        ws_scores['D1'] = "アスリートマインド"
        ws_scores['D1'].font = header_font
        ws_scores['D1'].fill = header_fill
        
        athlete_mind_data = [
            ['項目', 'スコア'],
            ['内省性', test_result.introspection or 0],
            ['自制心', test_result.self_control or 0],
            ['献身性', test_result.devotion or 0],
            ['直感性', test_result.intuition or 0],
            ['敏感性', test_result.sensitivity or 0],
            ['安定性', test_result.steadiness or 0],
            ['比較性', test_result.comparison or 0],
            ['結果志向', test_result.result or 0],
            ['主張性', test_result.assertion or 0],
            ['細密性', test_result.commitment or 0]
        ]
        
        for i, row in enumerate(athlete_mind_data):
            for j, value in enumerate(row):
                cell = ws_scores.cell(row=i+2, column=j+4, value=value)
                if i == 0:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
        
        # 自己肯定感
        ws_scores['G1'] = "自己肯定感"
        ws_scores['G1'].font = header_font
        ws_scores['G1'].fill = header_fill
        
        self_affirmation_data = [
            ['項目', 'スコア'],
            ['自己決定感', test_result.self_determination or 0],
            ['自己受容感', test_result.self_acceptance or 0],
            ['自己価値感', test_result.self_worth or 0],
            ['自己効力感', test_result.self_efficacy or 0]
        ]
        
        for i, row in enumerate(self_affirmation_data):
            for j, value in enumerate(row):
                cell = ws_scores.cell(row=i+2, column=j+7, value=value)
                if i == 0:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
        
        # グラフシート
        ws_charts = workbook.create_sheet("グラフ")
        
        # 列幅調整
        for ws in [ws_summary, ws_scores, ws_charts]:
            for column_cells in ws.columns:
                length = max(len(str(cell.value) or "") for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length + 2
        
        workbook.save(buffer)
        buffer.seek(0)
        
        return StreamingResponse(
            io.BytesIO(buffer.read()),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename=test_result_{result_id}.xlsx"}
        )
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Excel export failed: {str(e)}"
        )


@router.get("/csv/{result_id}")
async def export_csv(
    result_id: UUID,
    db: Session = Depends(get_db)
):
    """CSV形式でテスト結果をエクスポート"""
    try:
        # テスト結果を取得
        test_result = db.query(TestResult).filter(
            TestResult.result_id == result_id
        ).first()
        
        if not test_result:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Test result not found"
            )
        
        # CSVデータ準備
        output = io.StringIO()
        writer = csv.writer(output)
        
        # ヘッダー
        writer.writerow(['カテゴリ', '項目', 'スコア', '最大値'])
        
        # スポーツマンシップ
        sportsmanship_items = [
            ('勇気', test_result.courage or 0, 10),
            ('精神的強さ', test_result.resilience or 0, 10),
            ('協調性', test_result.cooperation or 0, 10),
            ('自然受容', test_result.natural_acceptance or 0, 10),
            ('非合理性', test_result.non_rationality or 0, 10)
        ]
        
        for item, score, max_score in sportsmanship_items:
            writer.writerow(['スポーツマンシップ', item, score, max_score])
        
        # アスリートマインド
        athlete_mind_items = [
            ('内省性', test_result.introspection or 0, 50),
            ('自制心', test_result.self_control or 0, 50),
            ('献身性', test_result.devotion or 0, 50),
            ('直感性', test_result.intuition or 0, 50),
            ('敏感性', test_result.sensitivity or 0, 50),
            ('安定性', test_result.steadiness or 0, 50),
            ('比較性', test_result.comparison or 0, 50),
            ('結果志向', test_result.result or 0, 50),
            ('主張性', test_result.assertion or 0, 50),
            ('細密性', test_result.commitment or 0, 50)
        ]
        
        for item, score, max_score in athlete_mind_items:
            writer.writerow(['アスリートマインド', item, score, max_score])
        
        # 自己肯定感
        self_affirmation_items = [
            ('自己決定感', test_result.self_determination or 0, 50),
            ('自己受容感', test_result.self_acceptance or 0, 50),
            ('自己価値感', test_result.self_worth or 0, 50),
            ('自己効力感', test_result.self_efficacy or 0, 50)
        ]
        
        for item, score, max_score in self_affirmation_items:
            writer.writerow(['自己肯定感', item, score, max_score])
        
        output.seek(0)
        
        return StreamingResponse(
            io.StringIO(output.getvalue()),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename=test_result_{result_id}.csv"}
        )
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"CSV export failed: {str(e)}"
        )


@router.get("/json/{result_id}")
async def export_json(
    result_id: UUID,
    db: Session = Depends(get_db)
):
    """JSON形式でテスト結果をエクスポート"""
    try:
        # テスト結果を取得
        test_result = db.query(TestResult).filter(
            TestResult.result_id == result_id
        ).first()
        
        if not test_result:
            raise HTTPException(
                status_code=status.HTTP_404_NOT_FOUND,
                detail="Test result not found"
            )
        
        # アスリートタイプ分析
        athlete_type = None
        try:
            athlete_mind_data = {
                'commitment': test_result.commitment or 0,
                'results_oriented': test_result.result or 0,
                'steady': test_result.steadiness or 0,
                'devoted': test_result.devotion or 0,
                'self_control': test_result.self_control or 0,
                'assertive': test_result.assertion or 0,
                'sensitive': test_result.sensitivity or 0,
                'intuitive': test_result.intuition or 0,
                'introspective': test_result.introspection or 0,
                'comparative': test_result.comparison or 0
            }
            athlete_type = analyze_athlete_type(athlete_mind_data, test_result.target_selection)
        except Exception as e:
            print(f"Athlete type analysis failed: {e}")
        
        # JSONデータ構築
        export_data = {
            "test_result": {
                "result_id": str(test_result.result_id),
                "target_selection": test_result.target_selection,
                "target_name": get_target_name(test_result.target_selection),
                "created_date": test_result.created_date.isoformat(),
                "scores": {
                    "sportsmanship": {
                        "courage": test_result.courage or 0,
                        "resilience": test_result.resilience or 0,
                        "cooperation": test_result.cooperation or 0,
                        "natural_acceptance": test_result.natural_acceptance or 0,
                        "non_rationality": test_result.non_rationality or 0
                    },
                    "athlete_mind": {
                        "introspection": test_result.introspection or 0,
                        "self_control": test_result.self_control or 0,
                        "devotion": test_result.devotion or 0,
                        "intuition": test_result.intuition or 0,
                        "sensitivity": test_result.sensitivity or 0,
                        "steadiness": test_result.steadiness or 0,
                        "comparison": test_result.comparison or 0,
                        "result": test_result.result or 0,
                        "assertion": test_result.assertion or 0,
                        "commitment": test_result.commitment or 0
                    },
                    "self_affirmation": {
                        "self_determination": test_result.self_determination or 0,
                        "self_acceptance": test_result.self_acceptance or 0,
                        "self_worth": test_result.self_worth or 0,
                        "self_efficacy": test_result.self_efficacy or 0
                    }
                },
                "athlete_type": athlete_type,
                "export_metadata": {
                    "export_date": datetime.now().isoformat(),
                    "version": "1.0.0",
                    "format": "JSON"
                }
            }
        }
        
        json_str = json.dumps(export_data, ensure_ascii=False, indent=2)
        
        return StreamingResponse(
            io.StringIO(json_str),
            media_type="application/json",
            headers={"Content-Disposition": f"attachment; filename=test_result_{result_id}.json"}
        )
        
    except Exception as e:
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"JSON export failed: {str(e)}"
        )


# ヘルパー関数
def get_target_name(target: str) -> str:
    """対象名を日本語に変換"""
    target_names = {
        "player": "選手",
        "coach": "指導者",
        "mother": "母親",
        "father": "父親",
        "adult": "大人・一般"
    }
    return target_names.get(target, target)


def get_grade(avg: float, max_score: float) -> str:
    """平均スコアから評価を算出"""
    percentage = (avg / max_score) * 100
    if percentage >= 80:
        return 'A'
    elif percentage >= 70:
        return 'B'
    elif percentage >= 60:
        return 'C'
    elif percentage >= 50:
        return 'D'
    else:
        return 'E'